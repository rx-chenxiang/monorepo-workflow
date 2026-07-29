from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


SKILL_ROOT = Path(__file__).resolve().parents[1]
GOLDEN = SKILL_ROOT / "examples" / "golden" / "customer_assignment.json"
HIGH_RISK_GOLDEN = SKILL_ROOT / "examples" / "golden" / "order_callback_high_risk.json"


def run_script(*args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, *args],
        cwd=SKILL_ROOT,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )


class TestYlScriptsTest(unittest.TestCase):
    def test_validate_golden_examples_pass_default_strict(self) -> None:
        for path in (GOLDEN, HIGH_RISK_GOLDEN):
            with self.subTest(path=path.name):
                result = run_script("scripts/validate_testcase_json.py", "--input", str(path))
                self.assertEqual(result.returncode, 0, result.stderr)
                self.assertIn("校验通过", result.stdout)

    def test_validate_rejects_non_contiguous_case_ids(self) -> None:
        data = json.loads(GOLDEN.read_text(encoding="utf-8"))
        data["test_cases"][1]["case_id"] = "TC-003"
        data["coverage_matrix"][1]["linked_case_ids"] = "TC-003"
        with tempfile.TemporaryDirectory() as tmp:
            invalid_path = Path(tmp) / "invalid.json"
            invalid_path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

            result = run_script("scripts/validate_testcase_json.py", "--input", str(invalid_path))

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("用例ID 编号必须", result.stderr)

    def test_workbook_contains_compatible_and_structured_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "cases.xlsx"
            result = run_script("scripts/write_testcase_workbook.py", "--input", str(GOLDEN), "--output", str(output))
            self.assertEqual(result.returncode, 0, result.stderr)

            workbook = load_workbook(output)
            self.assertEqual(
                workbook.sheetnames,
                ["测试用例", "结构化明细", "测试点覆盖矩阵", "二次核对报告"],
            )
            self.assertEqual(
                [cell.value for cell in workbook["测试用例"][1]],
                ["序号", "标题", "目录", "负责人", "前置条件", "步骤描述", "预期结果", "关联工作项", "优先级", "类型", "标签"],
            )
            self.assertEqual(
                [cell.value for cell in workbook["结构化明细"][1]],
                ["用例ID", "平台", "页面", "功能模块", "测试维度", "测试数据", "验证层级", "备注", "source_doc", "source_heading", "source_excerpt"],
            )
            self.assertIn("source_excerpt", [cell.value for cell in workbook["测试点覆盖矩阵"][1]])

    def test_markdown_mindmap_and_profile_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            markdown = tmp_path / "cases.md"
            mindmap = tmp_path / "mindmap.md"
            profile = tmp_path / "profile.md"

            md_result = run_script("scripts/write_testcase_markdown.py", "--input", str(GOLDEN), "--output", str(markdown))
            mindmap_result = run_script("scripts/write_test_design_mindmap.py", "--input", str(GOLDEN), "--output", str(mindmap))
            profile_result = run_script("scripts/update_business_profile.py", "--input", str(GOLDEN), "--profile", str(profile), "--source", "客户分配优化")

            self.assertEqual(md_result.returncode, 0, md_result.stderr)
            self.assertEqual(mindmap_result.returncode, 0, mindmap_result.stderr)
            self.assertEqual(profile_result.returncode, 0, profile_result.stderr)
            self.assertIn("来源文档", markdown.read_text(encoding="utf-8"))
            self.assertIn("mindmap", mindmap.read_text(encoding="utf-8"))
            self.assertIn("客户分配优化", profile.read_text(encoding="utf-8"))

    def test_preflight_checks_input_and_output_dir(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "需求.md"
            source.write_text("# 需求\n\n测试内容", encoding="utf-8")
            output_dir = Path(tmp) / "out"

            result = run_script("scripts/preflight.py", "--input", str(source), "--output-dir", str(output_dir))

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertIn("检查通过", result.stdout)


if __name__ == "__main__":
    unittest.main()
