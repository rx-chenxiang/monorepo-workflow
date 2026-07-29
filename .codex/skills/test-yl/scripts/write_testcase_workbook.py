#!/usr/bin/env python3
"""Write a three-sheet Chinese testcase workbook from JSON."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TEST_TYPE_VALUES = [
    "回归测试",
    "功能测试",
    "性能测试",
    "兼容性测试",
    "易用性测试",
    "安全性测试",
    "稳定性测试",
    "接口测试",
    "自动化测试",
    "安装部署测试",
    "冒烟测试",
]

TEST_DIMENSION_VALUES = [
    "功能入口",
    "UI展示",
    "交互效果",
    "更新机制",
    "连点测试",
    "断网测试",
    "正常场景",
    "异常场景",
    "旧版本兼容",
    "影响范围",
    "测试经验",
]

TEST_CASE_COLUMNS = [
    "序号",
    "标题",
    "目录",
    "负责人",
    "前置条件",
    "步骤描述",
    "预期结果",
    "关联工作项",
    "优先级",
    "类型",
    "标签",
]

STRUCTURED_DETAIL_COLUMNS = [
    "用例ID",
    "平台",
    "页面",
    "功能模块",
    "测试维度",
    "测试数据",
    "验证层级",
    "备注",
    "source_doc",
    "source_heading",
    "source_excerpt",
]

COVERAGE_COLUMNS = [
    "需求ID",
    "需求名称",
    "功能点",
    "数据表/接口",
    "测试点ID",
    "关联用例ID",
    "覆盖状态",
    "source_doc",
    "source_heading",
    "source_excerpt",
    "说明",
]

REVIEW_COLUMNS = ["核对维度", "核对结果", "说明"]

ALIASES = {
    "case_id": "用例ID",
    "id": "用例ID",
    "module": "目录",
    "platform": "平台",
    "page": "页面",
    "function_module": "功能模块",
    "test_dimension": "测试维度",
    "title": "标题",
    "precondition": "前置条件",
    "test_data": "测试数据",
    "steps": "步骤描述",
    "expected": "预期结果",
    "verification_layer": "验证层级",
    "work_item": "关联工作项",
    "linked_work_item": "关联工作项",
    "linked_work_items": "关联工作项",
    "requirement_id": "关联需求",
    "test_point_id": "关联测试点",
    "priority": "优先级",
    "type": "类型",
    "tags": "标签",
    "status": "状态",
    "owner": "负责人",
    "remark": "备注",
    "requirement_name": "需求名称",
    "feature": "功能点",
    "api_or_table": "数据表/接口",
    "linked_case_ids": "关联用例ID",
    "coverage_status": "覆盖状态",
    "dimension": "核对维度",
    "result": "核对结果",
    "description": "说明",
}


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in row.items():
        normalized[ALIASES.get(key, key)] = value
    return normalized


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def format_numbered(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(f"{index}. {item}" for index, item in enumerate(value, start=1))
    raw = str(value).strip()
    if not raw:
        return ""
    lines = [line.strip() for line in raw.splitlines() if line.strip()]
    if len(lines) > 1:
        numbered = [line for line in lines if re.match(r"^\s*\d+[\.、)]\s*", line)]
        if len(numbered) == len(lines):
            return "\n".join(lines)
        formatted = []
        for index, line in enumerate(lines, start=1):
            cleaned = re.sub(r"^[-*]\s*", "", line)
            formatted.append(f"{index}. {cleaned}")
        return "\n".join(formatted)
    return raw


def merge_nonempty(parts: list[Any], separator: str = " / ") -> str:
    values = [str(part).strip() for part in parts if str(part or "").strip()]
    return separator.join(values)


def enrich_testcase_row(row: dict[str, Any]) -> dict[str, Any]:
    row = dict(row)

    directory_detail = merge_nonempty([row.get("平台"), row.get("页面"), row.get("功能模块")])
    if not row.get("目录"):
        row["目录"] = directory_detail
    elif directory_detail and directory_detail not in str(row.get("目录")):
        row["目录"] = f"{row['目录']} / {directory_detail}"

    if not row.get("关联工作项"):
        row["关联工作项"] = merge_nonempty([row.get("关联需求"), row.get("关联测试点")], ",")

    tag_parts = [
        row.get("标签"),
        row.get("用例ID"),
        row.get("测试维度"),
        row.get("平台"),
        row.get("页面"),
        row.get("功能模块"),
    ]
    tag_values: list[str] = []
    for part in tag_parts:
        for item in re.split(r"[,，/]\s*", str(part or "")):
            item = item.strip()
            if item and item not in tag_values:
                tag_values.append(item)
    if tag_values:
        row["标签"] = ",".join(tag_values)

    return row


def append_sheet(workbook: Workbook, title: str, columns: list[str], rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(columns)

    for index, raw_row in enumerate(rows, start=1):
        row = normalize_row(raw_row)
        if title == "测试点覆盖矩阵":
            if "requirement_id" in raw_row:
                row["需求ID"] = raw_row["requirement_id"]
            if "test_point_id" in raw_row:
                row["测试点ID"] = raw_row["test_point_id"]
        if title == "测试用例":
            row = enrich_testcase_row(row)
        if "序号" in columns and not row.get("序号"):
            row["序号"] = index
        if "状态" in columns and not row.get("状态"):
            row["状态"] = "未执行"
        if "负责人" in columns and not row.get("负责人"):
            row["负责人"] = "系统生成"
        if "步骤描述" in row:
            row["步骤描述"] = format_numbered(row.get("步骤描述"))
        if "预期结果" in row:
            row["预期结果"] = format_numbered(row.get("预期结果"))
        sheet.append([stringify(row.get(column, "")) for column in columns])

    style_sheet(sheet, columns)


def style_sheet(sheet, columns: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_fill = PatternFill("solid", fgColor="F7FBFF")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = thin_fill

    width_map = {
        "序号": 8,
        "用例ID": 12,
        "目录": 18,
        "平台": 14,
        "页面": 22,
        "功能模块": 22,
        "测试维度": 16,
        "标题": 34,
        "前置条件": 32,
        "测试数据": 36,
        "步骤描述": 44,
        "预期结果": 44,
        "验证层级": 24,
        "关联需求": 16,
        "关联测试点": 18,
        "优先级": 10,
        "类型": 12,
        "标签": 20,
        "状态": 10,
        "负责人": 14,
        "备注": 32,
        "source_doc": 34,
        "source_heading": 28,
        "source_excerpt": 48,
        "需求ID": 14,
        "需求名称": 30,
        "功能点": 34,
        "数据表/接口": 28,
        "测试点ID": 14,
        "关联用例ID": 24,
        "覆盖状态": 14,
        "说明": 44,
        "核对维度": 22,
        "核对结果": 14,
    }

    for column_index, column_name in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width_map.get(column_name, 18)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    if sheet.title in {"测试用例", "结构化明细"} and "类型" in columns:
        type_column = get_column_letter(columns.index("类型") + 1)
        formula = '"' + ",".join(TEST_TYPE_VALUES) + '"'
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        validation.error = "类型只能从预设枚举中选择"
        validation.errorTitle = "类型无效"
        validation.prompt = "请选择测试类型"
        validation.promptTitle = "测试类型"
        sheet.add_data_validation(validation)
        validation.add(f"{type_column}2:{type_column}10000")

    if sheet.title in {"测试用例", "结构化明细"} and "测试维度" in columns:
        dimension_column = get_column_letter(columns.index("测试维度") + 1)
        formula = '"' + ",".join(TEST_DIMENSION_VALUES) + '"'
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        validation.error = "测试维度只能从预设枚举中选择"
        validation.errorTitle = "测试维度无效"
        validation.prompt = "请选择测试维度"
        validation.promptTitle = "测试维度"
        sheet.add_data_validation(validation)
        validation.add(f"{dimension_column}2:{dimension_column}10000")


def load_rows(data: dict[str, Any], key: str) -> list[dict[str, Any]]:
    rows = data.get(key, [])
    if not isinstance(rows, list):
        raise ValueError(f"{key} 必须是数组")
    return [row if isinstance(row, dict) else {"说明": row} for row in rows]


def main() -> int:
    parser = argparse.ArgumentParser(description="Write Chinese testcase workbook.")
    parser.add_argument("--input", required=True, help="Input JSON path.")
    parser.add_argument("--output", required=True, help="Output .xlsx path.")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser()
    output_path = Path(args.output).expanduser()
    data = json.loads(input_path.read_text(encoding="utf-8"))

    workbook = Workbook()
    workbook.remove(workbook.active)
    append_sheet(workbook, "测试用例", TEST_CASE_COLUMNS, load_rows(data, "test_cases"))
    append_sheet(workbook, "结构化明细", STRUCTURED_DETAIL_COLUMNS, load_rows(data, "test_cases"))
    append_sheet(workbook, "测试点覆盖矩阵", COVERAGE_COLUMNS, load_rows(data, "coverage_matrix"))
    append_sheet(workbook, "二次核对报告", REVIEW_COLUMNS, load_rows(data, "review_report"))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(str(output_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
